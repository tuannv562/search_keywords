from openpyxl import load_workbook
import urllib.parse
import requests
from copy import deepcopy
import time

INPUT_FILE_PATH = "/Users/vibbidi/Downloads/SEM_Keywords.xlsx"
SEARCH_API = "https://api.vibbidi.net/api/v6/graphql"

if __name__ == "__main__":
    print("Reading excel file...")
    wb = load_workbook(INPUT_FILE_PATH)
    ws = wb.active
    search_query_template = """{{search(query: "{}") {{tracks(startPoint: 0, itemsToGet: 50) {{id title}}}}}}"""
    result = []
    for row, cells in enumerate(ws.iter_rows(min_row=1, max_col=2)):
        if row == 0:
            continue
        if row % 50 == 0:
            print(f"Processed {row} keywords")
        row_data = []
        for col, cell in enumerate(cells):
            row_data.append(cell.value)
        encode = urllib.parse.quote(search_query_template.format(row_data[-1]))
        res = requests.get(f"https://api.vibbidi.net/api/v6/graphql?query={encode}")
        if res.status_code == 200:
            track_maps = res.json()["data"]["search"]["tracks"]
            if track_maps:
                for track_map in track_maps:
                    row_result = deepcopy(row_data)
                    row_result.append(track_map.get("id"))
                    row_result.append(track_map.get("title"))
                    result.append(row_result)
            else:
                row_data.append("")
                row_data.append("")
                result.append(row_data)
        else:
            row_data.append("")
            row_data.append("")
            result.append(row_data)
        time.sleep(1)
    ws_result = wb.create_sheet("Result")
    ws_result.append(["URL", "Keywords", "Track ID", "Track Title"])
    for row in result:
        ws_result.append(row)
    wb.save(INPUT_FILE_PATH)
    print("Finished. Please check result in excel file")
